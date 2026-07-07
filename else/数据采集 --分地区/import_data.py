#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SJG 数据批量导入脚本
功能：
  1. 读取 8 个 Excel 文件的数据采集汇总表
  2. 扫描素材库中所有媒体文件的路径
  3. 上传媒体文件到阿里云 OSS
  4. 生成带事务的 INSERT SQL（输出到 sjg_import.sql）

用法：
  第一步 — 仅生成 SQL（不上传 OSS，用于测试）：
    python3 import_data.py --sql-only

  第二步 — 上传 OSS + 生成 SQL：
    设置环境变量后运行：
    export OSS_ACCESS_KEY_ID=xxx
    export OSS_ACCESS_KEY_SECRET=xxx
    python3 import_data.py

依赖：
    pip3 install openpyxl oss2
"""

import openpyxl
import os
import sys
import re
import hashlib
from pathlib import Path
from datetime import datetime

# ============================================================
# 配置
# ============================================================
BASE_DIR = Path("/Users/a1/develop/vibecoding/sjg/数据采集 --分地区")
MEDIA_DIR = BASE_DIR / "黄河流域（山东段）文学景观素材库"
EXCEL_FILES = [
    "数据汇总-济南.xlsx",
    "数据汇总-泰安.xlsx",
    "数据汇总-德州.xlsx",
    "数据汇总-滨州.xlsx",
    "数据汇总-济宁.xlsx",
    "数据汇总--菏泽、聊城.xlsx",
    "数据汇总--淄博、东营.xlsx",
    "数据汇总 --黄河意象.xlsx",
]
OUTPUT_SQL = BASE_DIR / "sjg_import.sql"

# OSS 配置
OSS_BUCKET = "shandong-lit-landscape"
OSS_ENDPOINT = "https://oss-cn-beijing.aliyuncs.com"
OSS_BASE_URL = "https://shandong-lit-landscape.oss-cn-beijing.aliyuncs.com"

# 朝代名称 → ID 映射（对齐 sjg.sql 中的 dynasty 表）
DYNASTY_MAP = {
    "先秦": 1, "秦": 1,
    "西汉": 2, "东汉": 2, "汉": 2, "汉代": 2,
    "魏晋": 3, "西晋": 3, "东晋": 3, "南北朝": 3, "三国": 3,
    "唐": 4, "唐代": 4, "唐朝": 4,
    "宋": 5, "宋代": 5, "北宋": 5, "南宋": 5,
    "元": 6, "元代": 6,
    "明": 7, "明代": 7,
    "清": 8, "清代": 8, "清朝": 8,
    "金": 9, "金代": 9,
}

# 地区 → 区域简称（数据库 region 字段）
REGION_MAP = {
    "济南": "济南", "济南市": "济南", "济南市历下区": "济南", "济南市章丘区": "济南",
    "泰安": "泰安", "泰安市": "泰安", "泰安市泰山区": "泰安",
    "德州": "德州", "德州市": "德州",
    "滨州": "滨州", "滨州市": "滨州",
    "济宁": "济宁", "济宁市": "济宁", "济宁市曲阜市": "济宁",
    "聊城": "聊城", "聊城市": "聊城",
    "菏泽": "菏泽", "菏泽市": "菏泽",
    "淄博": "淄博", "淄博市": "淄博",
    "东营": "东营", "东营市": "东营",
    "青岛": "青岛", "青岛市": "青岛",
}


# ============================================================
# 工具函数
# ============================================================
def clean_filename(name):
    """清理文件名，去掉可能的多余扩展名和空格"""
    name = str(name).strip().strip('"').strip("'")
    # 修复 .jpg.jpg → .jpg, .mp4.mp4 → .mp4
    name = re.sub(r'\.(jpg|jpeg|png|webp|mp4|mp3)\1$', r'.\1', name, flags=re.IGNORECASE)
    # 修复 .jpg.jpeg → .jpeg, etc
    name = re.sub(r'\.(jpg|png)\.(jpeg|webp)$', r'.\2', name, flags=re.IGNORECASE)
    # 修复 .mp4 后的多余字符
    name = re.sub(r'\.mp4\..*$', '.mp4', name, flags=re.IGNORECASE)
    return name


def safe_str(val, default=""):
    if val is None:
        return default
    s = str(val).strip()
    return s if s and s.lower() != "none" else default


def parse_lnglat(val):
    """解析经纬度，返回 float 或 None"""
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def escape_sql(val):
    """MySQL 字符串转义"""
    if val is None:
        return "NULL"
    s = str(val)
    s = s.replace("\\", "\\\\")
    s = s.replace("'", "\\'")
    s = s.replace('"', '\\"')
    s = s.replace("\n", "\\n")
    return f"'{s}'"


# ============================================================
# 步骤 1: 构建媒体文件索引
# ============================================================
def build_media_index():
    """扫描素材库，建立 文件名 → 本地路径 的映射"""
    index = {}
    for filepath in MEDIA_DIR.rglob("*"):
        if filepath.is_file() and filepath.name != ".DS_Store":
            key = filepath.name.lower()
            if key not in index:
                index[key] = filepath
            else:
                # 同名文件，保留首次匹配（或根据父目录决定）
                pass
    return index


# ============================================================
# 步骤 2: 解析 Excel
# ============================================================
def parse_all_excels():
    """解析所有 Excel 的数据采集汇总 sheet，返回统一格式的记录列表"""
    all_rows = []

    for fname in EXCEL_FILES:
        fpath = BASE_DIR / fname
        if not fpath.exists():
            print(f"  ⚠ 文件不存在: {fname}")
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True)
        if "数据采集汇总" not in wb.sheetnames:
            print(f"  ⚠ {fname} 没有'数据采集汇总' sheet")
            continue

        ws = wb["数据采集汇总"]
        headers = [safe_str(c.value, f"COL_{i}") for i, c in enumerate(ws[2])]

        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            spot_name = safe_str(row[0])
            if not spot_name:
                continue

            record = {}
            for i, h in enumerate(headers):
                record[h] = safe_str(row[i]) if i < len(row) else ""

            record["_source"] = fname
            record["_row"] = row_idx
            record["景观标准名 *"] = spot_name
            all_rows.append(record)
            count += 1

        print(f"  ✓ {fname}: {count} 条记录")

    print(f"\n  总计: {len(all_rows)} 条数据记录")
    return all_rows


# ============================================================
# 步骤 3: 数据去重与实体构建
# ============================================================
def build_entities(all_rows):
    """从 Excel 记录中提取去重后的景点、诗人、诗词、事件"""

    # --- 朝代 (直接使用硬编码映射) ---
    dynasties = {}  # name → id

    # --- 景点去重 ---
    spots = {}  # name → {fields}
    for r in all_rows:
        name = r["景观标准名 *"]
        if name not in spots:
            lng = parse_lnglat(r.get("经度 *", ""))
            lat = parse_lnglat(r.get("纬度 *", ""))
            region = extract_region(r.get("所在市县", ""), name)

            spots[name] = {
                "name": name,
                "type_main": r.get("景观类型（大类）", ""),
                "type_sub": r.get("景观类型（子类）", ""),
                "address": r.get("具体位置", "") or region,
                "longitude": lng,
                "latitude": lat,
                "river_section": r.get("对应黄河河道", ""),
                "history": r.get("历史沿革", ""),
                "culture": r.get("文化内涵", ""),
                "region": region,
                "image_file": clean_filename(r.get("景观实景图名", "")),
                "video_file": clean_filename(r.get("景观视频名", "")),
                "images": [],
                "videos": [],
            }
        # 收集该景点的所有媒体文件
        img = clean_filename(r.get("景观实景图名", ""))
        vid = clean_filename(r.get("景观视频名", ""))
        if img and img not in spots[name]["images"]:
            spots[name]["images"].append(img)
        if vid and vid not in spots[name]["videos"]:
            spots[name]["videos"].append(vid)

    # --- 诗人去重 ---
    poets = {}  # name → {fields}
    for r in all_rows:
        author = r.get("作者 *", "")
        if not author:
            continue
        if author not in poets:
            dynasty_name = r.get("朝代 *", "")
            dynasty_id = DYNASTY_MAP.get(dynasty_name, None)
            poets[author] = {
                "name": author,
                "dynasty_id": dynasty_id,
                "dynasty_name": dynasty_name,
            }

    # --- 诗词 ---
    poems = []  # list of {fields, spot_name, poet_name}
    for r in all_rows:
        title = r.get("诗题 *", "")
        if not title:
            continue
        author = r.get("作者 *", "")
        dynasty_name = r.get("朝代 *", "")
        content_raw = r.get("诗词正文 *", "")

        # 诗词正文中的 / 换行可能有多种表示
        content = content_raw.replace("/", "\n").replace("\\n", "\n").strip()

        poem = {
            "title": title,
            "poet_name": author,
            "spot_name": r["景观标准名 *"],
            "dynasty_name": dynasty_name,
            "dynasty_id": DYNASTY_MAP.get(dynasty_name, None),
            "content": content,
            "imagery": r.get("核心意象", ""),
            "source": r.get("文献出处", ""),
            "background": r.get("文化内涵", ""),
            "era": r.get("创作时期", ""),
            "annotation": r.get("备注", ""),
            "ai_video_file": clean_filename(r.get("AI视频名", "")),
            "audio_file": clean_filename(r.get("名家朗读音频名", "")),
        }
        poems.append(poem)

    # --- 事件 ---
    events = {}  # title → {fields}
    for r in all_rows:
        title = r.get("诗题 *", "")
        # 事件目前 Excel 中没有专门的列，用文化内涵和备注推断
        # 保持现有 sjg.sql 中的事件不变

    return spots, poets, poems


def extract_region(city_field, spot_name):
    """从城市字段提取区域简称"""
    if not city_field:
        return "其他"
    city_field = str(city_field).strip()
    for key, value in REGION_MAP.items():
        if key in city_field:
            return value
    # 从景点名推断
    for key, value in REGION_MAP.items():
        if key in spot_name:
            return value
    return city_field.split("市")[0].split("区")[0] if city_field else "其他"


# ============================================================
# 步骤 4: 生成 OSS 上传清单
# ============================================================
def classify_media(spots, poets, poems, media_index):
    """
    将所有媒体文件分类到 OSS 目录结构：
      poets/      — 作者照片/画像
      spots/      — 景点实景图
      video/      — 视频
      audio/      — 音频
      events/     — 事件图片
    """
    uploads = []  # [(local_path, oss_key)]

    # 扫描媒体目录中所有文件，按命名规则分类
    for filename, filepath in media_index.items():
        fn = str(filepath.name).lower()
        rel_path = str(filepath.relative_to(MEDIA_DIR))

        if "作者" in rel_path or "画像" in rel_path:
            oss_dir = "poets"
        elif "视频" in rel_path or fn.endswith(".mp4"):
            oss_dir = "video"
        elif fn.endswith(".mp3"):
            oss_dir = "audio"
        else:
            oss_dir = "spots"

        # OSS 路径：去除中文子目录，直接扁平化
        oss_key = f"{oss_dir}/{filepath.name}"
        uploads.append((str(filepath), oss_key))

    # 对 Excel 中引用的文件名也做映射
    excel_files = set()
    for s in spots.values():
        for f in s["images"] + s["videos"]:
            if f:
                excel_files.add(f.lower())
    for p in poems:
        for f in [p["ai_video_file"], p["audio_file"]]:
            if f:
                excel_files.add(f.lower())

    return uploads, excel_files


# ============================================================
# 步骤 5: OSS 上传
# ============================================================
def upload_to_oss(uploads):
    """批量上传文件到 OSS"""
    try:
        import oss2
    except ImportError:
        print("  ❌ 请先安装 oss2: pip3 install oss2")
        return {}

    access_key_id = os.environ.get("OSS_ACCESS_KEY_ID", "")
    access_key_secret = os.environ.get("OSS_ACCESS_KEY_SECRET", "")

    if not access_key_id or not access_key_secret:
        print("  ❌ 请设置环境变量 OSS_ACCESS_KEY_ID 和 OSS_ACCESS_KEY_SECRET")
        print("     export OSS_ACCESS_KEY_ID=xxx")
        print("     export OSS_ACCESS_KEY_SECRET=xxx")
        return {}

    auth = oss2.Auth(access_key_id, access_key_secret)
    bucket = oss2.Bucket(auth, OSS_ENDPOINT, OSS_BUCKET)

    url_map = {}  # local_path → oss_url
    success = 0
    fail = 0

    for local_path, oss_key in uploads:
        try:
            bucket.put_object_from_file(oss_key, local_path)
            url = f"{OSS_BASE_URL}/{oss_key}"
            url_map[local_path] = url
            success += 1
            if success % 20 == 0:
                print(f"    已上传 {success}/{len(uploads)} ...")
        except Exception as e:
            print(f"    ⚠ 上传失败 [{oss_key}]: {e}")
            fail += 1

    print(f"  ✓ 上传完成: 成功 {success}, 失败 {fail}")
    return url_map


# ============================================================
# 步骤 6: 生成带事务的 SQL
# ============================================================
def generate_sql(spots, poets, poems, media_index, url_map):
    """生成完整的 INSERT SQL，包含事务和外键关联"""

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def get_oss_url(filename, category="spots"):
        """根据文件名查找对应的 OSS URL"""
        if not filename:
            return None
        fn = filename.lower()
        # 在 media_index 中查找匹配文件
        for mf, mp in media_index.items():
            if mf == fn:
                # 如果已上传，返回 OSS URL
                if str(mp) in url_map:
                    return url_map[str(mp)]
                # 否则生成预期的 OSS URL
                if "作者" in str(mp) or "画像" in str(mp):
                    return f"{OSS_BASE_URL}/poets/{mp.name}"
                if mp.suffix.lower() in (".mp4",):
                    return f"{OSS_BASE_URL}/video/{mp.name}"
                if mp.suffix.lower() in (".mp3",):
                    return f"{OSS_BASE_URL}/audio/{mp.name}"
                return f"{OSS_BASE_URL}/spots/{mp.name}"
        return None

    sql = []
    sql.append("-- ========================================")
    sql.append(f"-- SJG 数据批量导入 SQL (事务包裹)")
    sql.append(f"-- 生成时间: {now}")
    sql.append(f"-- 景点数: {len(spots)}, 诗人: {len(poets)}, 诗词: {len(poems)}")
    sql.append("-- ========================================")
    sql.append("")
    sql.append("START TRANSACTION;")
    sql.append("")

    # 设置外键检查临时关闭，避免插入顺序问题
    sql.append("SET FOREIGN_KEY_CHECKS = 0;")
    sql.append("")

    # --- 清表（可选，取消注释来清空后重新导入）---
    # sql.append("DELETE FROM poem_event;")
    # sql.append("DELETE FROM poem;")
    # sql.append("DELETE FROM event;")
    # sql.append("DELETE FROM scenic_spot;")
    # sql.append("DELETE FROM poet;")
    # sql.append("")

    # --- 1. 插入朝代 (如果缺少则补) ---
    existing_dynasties = set(DYNASTY_MAP.values())
    sql.append("-- ===== 朝代 ===== --")
    sql.append("INSERT IGNORE INTO dynasty (id, name) VALUES")
    dynasty_vals = []
    for dname, did in sorted(DYNASTY_MAP.items(), key=lambda x: x[1]):
        if did <= 8:  # 只处理 1-8（对齐已有 SQL）
            continue
        dynasty_vals.append(f"({did}, {escape_sql(dname)})")
    if dynasty_vals:
        sql.append(",\n".join(dynasty_vals) + ";")
    else:
        sql.append("-- (朝代已全部存在)")
    sql.append("")

    # --- 2. 插入诗人 ---
    sql.append("-- ===== 诗人 ===== --")
    poet_id_map = {}  # name → id
    poet_inserts = []
    for pid, (name, data) in enumerate(poets.items(), start=1):
        poet_id_map[name] = pid
        did = data["dynasty_id"] if data["dynasty_id"] and 1 <= data["dynasty_id"] <= 8 else "NULL"
        # 查找诗人头像 URL
        avatar_url = None
        for mf, mp in media_index.items():
            # 按诗人名字匹配作者图片
            rel_str = str(mp).lower()
            if name.lower() in rel_str and ("作者" in rel_str or "画像" in rel_str):
                avatar_url = get_oss_url(mf) or f"{OSS_BASE_URL}/poets/{mp.name}"
                break

        avatar = escape_sql(avatar_url) if avatar_url else "NULL"
        poet_inserts.append(
            f"({pid}, {escape_sql(name)}, {did}, "
            f"NULL, NULL, NULL, NULL, "
            f"{avatar}, NULL, NULL, {escape_sql(now)}, {escape_sql(now)})"
        )

    sql.append(
        "INSERT INTO poet (id, name, dynasty_id, birth_year, death_year, birthplace, biography, avatar_url, avatar_anime_url, style, created_at, updated_at) VALUES"
    )
    sql.append(",\n".join(poet_inserts) + ";")
    sql.append("")

    # --- 3. 插入景点 ---
    sql.append("-- ===== 景点 ===== --")
    spot_id_map = {}  # name → id
    spot_inserts = []
    for sid, (name, data) in enumerate(spots.items(), start=1):
        spot_id_map[name] = sid
        lng = data["longitude"] if data["longitude"] else "NULL"
        lat = data["latitude"] if data["latitude"] else "NULL"

        # 查找匹配的图片文件
        image_url = None
        anime_url = None
        for mf, mp in media_index.items():
            rel_str = str(mp).lower()
            if name.lower() in rel_str and "实景" in rel_str:
                url = get_oss_url(mf) or f"{OSS_BASE_URL}/spots/{mp.name}"
                if not image_url:
                    image_url = url

        img = escape_sql(image_url) if image_url else "NULL"
        ani = escape_sql(anime_url) if anime_url else "NULL"

        # description 从 history 和 culture 拼接
        desc_parts = []
        if data.get("history"):
            desc_parts.append(data["history"])
        if data.get("culture"):
            desc_parts.append(data["culture"])
        desc = escape_sql("；".join(desc_parts) if desc_parts else data["name"])

        spot_inserts.append(
            f"({sid}, {escape_sql(name)}, {desc}, {lng}, {lat}, "
            f"{escape_sql(data['address'])}, {img}, {ani}, {escape_sql(data['region'])}, "
            f"{escape_sql(now)}, {escape_sql(now)})"
        )

    sql.append(
        "INSERT INTO scenic_spot (id, name, description, longitude, latitude, address, image_url, image_anime_url, region, created_at, updated_at) VALUES"
    )
    sql.append(",\n".join(spot_inserts) + ";")
    sql.append("")

    # --- 4. 插入诗词 ---
    sql.append("-- ===== 诗词 ===== --")
    poem_inserts = []
    poem_id = 0
    for p in poems:
        poem_id += 1
        pid = poet_id_map.get(p["poet_name"], "NULL")
        sid = spot_id_map.get(p["spot_name"], "NULL")
        did = p["dynasty_id"] if p["dynasty_id"] and 1 <= p["dynasty_id"] <= 8 else "NULL"

        # 查找媒体 URL
        audio_url = None
        video_url = None
        for mf, mp in media_index.items():
            fn = str(mp.name).lower()
            # 按诗名模糊匹配
            shortened = p["title"][:4].lower()
            if shortened in fn or p["title"].lower()[:6] in fn:
                url = get_oss_url(mf) or f"{OSS_BASE_URL}/audio/{mp.name}"
                if mp.suffix.lower() == ".mp4":
                    video_url = url
                elif mp.suffix.lower() == ".mp3":
                    audio_url = url

        aud = escape_sql(audio_url) if audio_url else "NULL"
        vid = escape_sql(video_url) if video_url else "NULL"

        # sentiment_tags → JSON 数组
        imagery = p["imagery"]
        tags = []
        if imagery:
            tags = [t.strip() for t in imagery.replace("、", ",").replace("，", ",").split(",") if t.strip()]

        poem_inserts.append(
            f"({poem_id}, {escape_sql(p['title'])}, {escape_sql(p['content'])}, "
            f"{pid}, {did}, {sid}, "
            f"{escape_sql(p['annotation'])}, {escape_sql(p['background'])}, "
            f"{aud}, {vid}, {escape_sql(str(tags))}, "
            f"{escape_sql(now)}, {escape_sql(now)})"
        )

    sql.append(
        "INSERT INTO poem (id, title, content, poet_id, dynasty_id, spot_id, annotation, background, audio_url, video_url, sentiment_tags, created_at, updated_at) VALUES"
    )
    sql.append(",\n".join(poem_inserts) + ";")
    sql.append("")

    # --- 5. 自动更新主键自增值 ---
    sql.append("-- ===== 更新自增 ID ===== --")
    max_poet = len(poets)
    max_spot = len(spots)
    sql.append(f"ALTER TABLE poet AUTO_INCREMENT = {max_poet + 1};")
    sql.append(f"ALTER TABLE scenic_spot AUTO_INCREMENT = {max_spot + 1};")
    sql.append(f"ALTER TABLE poem AUTO_INCREMENT = {poem_id + 1};")
    sql.append("")

    # --- 恢复外键 ---
    sql.append("SET FOREIGN_KEY_CHECKS = 1;")
    sql.append("")

    # --- 提交事务 ---
    sql.append("COMMIT;")
    sql.append("")
    sql.append("-- ===== 导入完成 =====")

    return "\n".join(sql)


# ============================================================
# 主流程
# ============================================================
def main():
    sql_only = "--sql-only" in sys.argv

    print("=" * 60)
    print("  SJG 数据批量导入工具")
    print("=" * 60)

    # 步骤 1: 构建媒体索引
    print("\n[1/5] 扫描媒体文件...")
    media_index = build_media_index()
    print(f"  ✓ 找到 {len(media_index)} 个媒体文件")

    # 步骤 2: 解析 Excel
    print("\n[2/5] 解析 Excel 数据...")
    all_rows = parse_all_excels()

    # 步骤 3: 构建实体
    print("\n[3/5] 数据去重与实体构建...")
    spots, poets, poems = build_entities(all_rows)
    print(f"  ✓ 去重后: {len(spots)} 个景点, {len(poets)} 位诗人, {len(poems)} 首诗词")

    # 打印实体清单
    print(f"\n  📍 景点清单 ({len(spots)}):")
    for name in sorted(spots.keys()):
        print(f"     - {name} ({spots[name]['region']})")

    print(f"\n  👤 诗人清单 ({len(poets)}):")
    for name, data in sorted(poets.items()):
        did = data["dynasty_name"] or "未知"
        print(f"     - {name} [{did}]")

    # 步骤 4: 分类媒体 & 上传清单
    print("\n[4/5] 生成 OSS 上传清单...")
    uploads, excel_files = classify_media(spots, poets, poems, media_index)
    print(f"  ✓ 待上传文件: {len(uploads)} 个")

    # 步骤 5: OSS 上传 (非 --sql-only 模式)
    url_map = {}
    if not sql_only:
        print("\n[5/5] 上传到 OSS...")
        url_map = upload_to_oss(uploads)
    else:
        print("\n[5/5] --sql-only 模式，跳过 OSS 上传")
        print("  OSS 文件列表:")
        for local_path, oss_key in uploads[:10]:
            print(f"    {oss_key} ← {Path(local_path).name}")
        if len(uploads) > 10:
            print(f"    ... 共 {len(uploads)} 个文件")

    # 步骤 6: 生成 SQL
    print("\n生成 SQL 导入脚本...")
    sql_content = generate_sql(spots, poets, poems, media_index, url_map)

    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write(sql_content)

    print(f"  ✓ SQL 已生成: {OUTPUT_SQL}")
    print(f"  大小: {len(sql_content):,} 字符, ~{sql_content.count(chr(10))} 行")
    print()
    print("=" * 60)
    print("  下一步:")
    print("    1. 检查生成的 SQL（景点/诗人数量和图片 URL 是否正确）")
    print("    2. 在正式导入前，建议先备份数据库: mysqldump -u root -p sjg > sjg_backup.sql")
    print(f"   3. 导入: mysql -u root -p sjg < {OUTPUT_SQL}")
    print("=" * 60)


if __name__ == "__main__":
    main()
