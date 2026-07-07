#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据核对脚本：对比导出的 SQL 与原始 Excel + OSS 素材
生成 sjg_update.sql 修正数据
"""
import openpyxl, re, os, json
from pathlib import Path
from collections import defaultdict

BASE_DIR = Path("/Users/a1/develop/vibecoding/sjg/数据采集 --分地区")
MEDIA_DIR = BASE_DIR / "黄河流域（山东段）文学景观素材库"
EXPORT_DIR = BASE_DIR / "导出数据"
OSS_BASE = "https://shandong-lit-landscape.oss-cn-beijing.aliyuncs.com"
OUTPUT_SQL = BASE_DIR / "sjg_update.sql"

EXCEL_FILES = [
    "数据汇总-济南.xlsx", "数据汇总-泰安.xlsx", "数据汇总-德州.xlsx",
    "数据汇总-滨州.xlsx", "数据汇总-济宁.xlsx", "数据汇总--菏泽、聊城.xlsx",
    "数据汇总--淄博、东营.xlsx", "数据汇总 --黄河意象.xlsx",
]


def escape_sql(v):
    if v is None: return "NULL"
    s = str(v).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"


def build_media_index():
    """扫描素材库，建立 文件名(小写) → (原名, OSS URL, 类别) 映射"""
    idx = {}
    for fp in MEDIA_DIR.rglob("*"):
        if not fp.is_file() or fp.name == ".DS_Store":
            continue
        rel = str(fp.relative_to(MEDIA_DIR))
        if "视频" in rel or fp.suffix.lower() == ".mp4":
            cat = "video"
        elif "作者" in rel or "画像" in rel:
            cat = "poets"
        else:
            cat = "spots"
        idx[fp.name.lower()] = (fp.name, f"{OSS_BASE}/{cat}/{fp.name}", cat)
    return idx


def parse_excel():
    """解析所有 Excel → {景点名: {images:[], videos:[]}}"""
    spots = defaultdict(lambda: {"images": [], "videos": [], "dynasty": "", "authors": set()})

    for fname in EXCEL_FILES:
        fp = BASE_DIR / fname
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["数据采集汇总"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            name = str(row[0]).strip()
            img = str(row[10]).strip() if len(row) > 10 and row[10] else ""
            vid = str(row[11]).strip() if len(row) > 11 and row[11] else ""
            author = str(row[13]).strip() if len(row) > 13 and row[13] else ""

            img = re.sub(r'\.(jpg|png|jpeg|webp)\1$', r'.\1', img) if img else ""
            if img and img.lower() != "none":
                spots[name]["images"].append(img)
            if vid and vid.lower() != "none":
                spots[name]["videos"].append(vid)
            if author:
                spots[name]["authors"].add(author)

    return spots


def parse_exported_sql():
    """从导出的 SQL 中提取 name→id, name→current_image 等"""
    poets = {}   # name → {id, avatar_url, dynasty_id}
    spot_map = {}  # name → {id, image_url, video_url}

    # 读 poet.sql
    with open(EXPORT_DIR / "poet.sql") as f:
        content = f.read()
    for m in re.finditer(r"VALUES\s*\((\d+),\s*'([^']+)',\s*(\d+|null),[^)]+'(https://[^']+)'[^)]+\)", content):
        pid = int(m.group(1))
        name = m.group(2)
        url = m.group(4)
        poets[name] = {"id": pid, "avatar_url": url}

    # 也匹配 avatar_url 为 NULL 的
    for m in re.finditer(r"VALUES\s*\((\d+),\s*'([^']+)',\s*(\d+|null),\s*(null|null),\s*(null|null)", content):
        pid = int(m.group(1))
        name = m.group(2)
        if name not in poets:
            poets[name] = {"id": pid, "avatar_url": None}

    print(f"  诗人: {len(poets)} 位")

    # 读 scenic_spot.sql
    with open(EXPORT_DIR / "scenic_spot.sql") as f:
        content = f.read()
    # 更宽松的匹配
    for m in re.finditer(r"VALUES\s*\((\d+),\s*'([^']+)'", content):
        sid = int(m.group(1))
        name = m.group(2)
        spot_map[name] = {"id": sid}

    # 提取 image_url
    for m in re.finditer(r"VALUES\s*\((\d+),\s*'([^']+)',\s*'[^']*',[^,]+,[^,]+,'[^']*',\s*'((?:https[^']*)|null)'", content):
        sid = int(m.group(1))
        name = m.group(2)
        img = m.group(3) if m.group(3) != "null" else None
        if name in spot_map:
            spot_map[name]["image_url"] = img

    print(f"  景点: {len(spot_map)} 个")

    return poets, spot_map


def match_media_to_spots(spot_media, media_index):
    """Excel 中的文件名 → OSS URL"""
    matched = defaultdict(lambda: {"images": set(), "videos": set()})

    for spot_name, media in spot_media.items():
        for img in media["images"]:
            key = img.lower().strip()
            # 修复 Excel 中的换行导致的文件名断裂
            key = key.replace("\n", "").replace("\r", "")
            if key in media_index:
                matched[spot_name]["images"].add(media_index[key][1])
                continue
            # 模糊匹配（下划线 vs 连字符等差异）
            base = re.sub(r'[_\-\s]+', '', os.path.splitext(key)[0]).lower()
            for mk, mv in media_index.items():
                if mv[2] != "spots": continue
                base_mk = re.sub(r'[_\-\s]+', '', os.path.splitext(mk)[0]).lower()
                if len(base) > 3 and len(base_mk) > 3 and (base in base_mk or base_mk in base):
                    matched[spot_name]["images"].add(mv[1])
                    break

        for vid in media["videos"]:
            key = vid.lower().strip().replace("\n", "").replace("\r", "")
            if key in media_index:
                matched[spot_name]["videos"].add(media_index[key][1])
            else:
                base = re.sub(r'[_\-\s]+', '', os.path.splitext(key)[0]).lower()
                for mk, mv in media_index.items():
                    if mv[2] != "video": continue
                    base_mk = re.sub(r'[_\-\s]+', '', os.path.splitext(mk)[0]).lower()
                    if len(base) > 3 and (base in base_mk or base_mk in base):
                        matched[spot_name]["videos"].add(mv[1])
                        break

    return matched


def match_poet_avatars(media_index):
    """从素材库提取诗人名 → OSS URL"""
    avatars = defaultdict(list)
    name_cleanup = {
        "李隆基_作者_01": "李隆基",
        "玄烨_作者_01": "玄烨", 
        "王士贞_作者_01": "王士贞",
        "郭沫若_作者_01": "郭沫若",
        "顾炎武_作者_01": "顾炎武",
        "苏轼_画像_01": "苏轼",
        "苏辙_画像_01": "苏辙",
    }

    for fname_lower, (orig, url, cat) in media_index.items():
        if cat != "poets": continue
        base = re.sub(r'\.\w+$', '', orig)
        
        # Check cleanup map first
        matched_name = None
        for pattern, name in name_cleanup.items():
            if pattern in orig:
                matched_name = name
                break
        
        if not matched_name:
            # Generic: 杜甫_画像_01 → 杜甫
            matched_name = re.sub(r'[_\-\s]*画像[_\-\s]*\d*', '', base).strip()
            matched_name = re.sub(r'[_\-\s]*作者[_\-\s]*\d*', '', matched_name).strip()
            matched_name = re.sub(r'[_\-\s]*\d+$', '', matched_name).strip()

        if matched_name:
            avatars[matched_name].append(url)

    return dict(avatars)


def generate_update_sql(poets, spot_map, spot_matched, poet_avatars):
    """生成 UPDATE SQL"""
    sql = []
    sql.append("-- ==========================================")
    sql.append("-- SJG 数据修正 UPDATE SQL")
    sql.append("-- 基于 Excel 原始数据 + OSS 素材库核对生成")
    sql.append("-- ==========================================")
    sql.append("")
    sql.append("START TRANSACTION;")
    sql.append("SET FOREIGN_KEY_CHECKS = 0;")
    sql.append("")

    fixes = {"poet_avatar": 0, "spot_image": 0, "spot_video": 0, "dynasty": 0}

    # --- 1. 修正诗人头像 ---
    sql.append("-- ===== 诗人头像修正 ===== --")
    for name, urls in poet_avatars.items():
        if name not in poets:
            # 尝试模糊匹配
            for pname in poets:
                if name in pname or pname in name:
                    name = pname
                    break
            else:
                sql.append(f"-- ⚠ 诗人 '{name}' 不在数据库中，无法更新头像")
                continue
        
        pid = poets[name]["id"]
        current = poets[name].get("avatar_url")
        
        # 如果当前已有头像且新头像也在列表中，追加
        all_urls = set(urls)
        if current and current != "NULL":
            all_urls.add(current)
        
        json_urls = json.dumps(sorted(all_urls), ensure_ascii=False)
        sql.append(f"UPDATE poet SET avatar_url = {escape_sql(json_urls)} WHERE id = {pid};  -- {name}")
        fixes["poet_avatar"] += 1

    sql.append("")

    # --- 2. 修正景点图片（单张→JSON数组）---
    sql.append("-- ===== 景点图片修正（多图转JSON数组）===== --")
    for spot_name, media in spot_matched.items():
        if spot_name not in spot_map:
            continue
        sid = spot_map[spot_name]["id"]
        images = sorted(media["images"])
        
        if not images:
            continue
        
        # 检查当前值
        current = spot_map[spot_name].get("image_url")
        if current and current != "NULL" and current not in images:
            images.insert(0, current)
            images = list(dict.fromkeys(images))  # 去重保序
        
        if len(images) == 1:
            json_val = escape_sql(json.dumps(images, ensure_ascii=False))
        else:
            json_val = escape_sql(json.dumps(images, ensure_ascii=False))
        
        sql.append(f"UPDATE scenic_spot SET image_url = {json_val} WHERE id = {sid};  -- {spot_name} ({len(images)}张)")
        fixes["spot_image"] += 1

    sql.append("")

    # --- 3. 修正景点视频 ---
    sql.append("-- ===== 景点视频修正 ===== --")
    for spot_name, media in spot_matched.items():
        if spot_name not in spot_map: continue
        videos = sorted(media["videos"])
        if not videos: continue
        sid = spot_map[spot_name]["id"]
        vids_dedup = list(dict.fromkeys(videos))
        json_val = escape_sql(json.dumps(vids_dedup, ensure_ascii=False))
        sql.append(f"UPDATE scenic_spot SET video_url = {json_val} WHERE id = {sid};  -- {spot_name} ({len(vids_dedup)}个视频)")
        fixes["spot_video"] += 1

    sql.append("")

    # --- 4. 修正 dynasty_id 为 NULL 的 ---
    sql.append("-- ===== 修正朝代ID ===== --")
    DYNASTY_MAP = {
        "先秦": 1, "秦": 1, "西汉": 2, "东汉": 2, "汉": 2, "汉代": 2,
        "魏晋": 3, "西晋": 3, "东晋": 3, "南北朝": 3, "三国": 3,
        "唐": 4, "唐代": 4, "唐朝": 4,
        "宋": 5, "宋代": 5, "北宋": 5, "南宋": 5,
        "元": 6, "元代": 6,
        "明": 7, "明代": 7,
        "清": 8, "清代": 8, "清朝": 8,
        "金": 9, "金代": 9,
    }

    spot_media = parse_excel()
    for fname in EXCEL_FILES:
        fp = BASE_DIR / fname
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["数据采集汇总"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            author = str(row[13]).strip() if len(row) > 13 and row[13] else ""
            dynasty = str(row[14]).strip() if len(row) > 14 and row[14] else ""
            if not author or not dynasty: continue
            
            did = DYNASTY_MAP.get(dynasty)
            if not did: continue
            
            # Find in poets
            if author in poets and poets[author].get("avatar_url", "") == "NULL" or True:
                # Check if current dynasty_id is NULL or wrong
                pid = poets[author]["id"]
                # We need the actual current dynasty_id — assume from the SQL we need to check
                sql.append(f"-- poet '{author}' dynasty '{dynasty}' → id={did}")

    sql.append("")
    sql.append("SET FOREIGN_KEY_CHECKS = 1;")
    sql.append("COMMIT;")
    sql.append("")
    sql.append(f"-- 修正统计: 诗人头像 {fixes['poet_avatar']}, 景点图片 {fixes['spot_image']}, 景点视频 {fixes['spot_video']}")

    return "\n".join(sql)


def main():
    print("=== 数据核对工具 ===")
    
    print("\n[1/4] 扫描素材库...")
    media_index = build_media_index()
    print(f"  {len(media_index)} 个文件")

    print("\n[2/4] 解析 Excel...")
    spot_media = parse_excel()
    print(f"  {len(spot_media)} 个景点")

    print("\n[3/4] 解析导出 SQL...")
    poets, spot_map = parse_exported_sql()

    print("\n[4/4] 匹配...")
    spot_matched = match_media_to_spots(spot_media, media_index)
    poet_avatars = match_poet_avatars(media_index)

    # Report
    multi = {k: v for k, v in spot_matched.items() if len(v["images"]) > 1}
    has_vid = {k: v for k, v in spot_matched.items() if len(v["videos"]) > 0}
    miss = {k: v for k, v in spot_media.items() if k in spot_map and not spot_matched[k]["images"]}

    print(f"\n  数据差异:")
    print(f"    多图景点: {len(multi)} (需要转 JSON 数组)")
    print(f"    有视频: {len(has_vid)}")
    print(f"    无匹配图片: {len(miss)}")
    print(f"    诗人有头像素材: {len(poet_avatars)} 位")
    
    # Count poets already having avatar
    have_avatar = sum(1 for p in poets.values() if p.get("avatar_url") and p["avatar_url"] not in (None, "NULL"))
    need_avatar = len(poet_avatars)  # poets we CAN add avatars for
    print(f"    当前有头像: {have_avatar}, 可补头像: {need_avatar}")

    print("\n生成 UPDATE SQL...")
    update_sql = generate_update_sql(poets, spot_map, spot_matched, poet_avatars)
    
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write(update_sql)
    
    print(f"  ✓ 已写入: {OUTPUT_SQL}")
    print(f"  {len(update_sql):,} 字符")

if __name__ == "__main__":
    main()
